from openpyxl import load_workbook
import random

fisier=load_workbook("D:\\Scoala\\an3\\PYTHON_IA\\PYTHON_LAB6\\Ex.xlsx")
fila=fisier["PAG1"]

c1=[]
c2=[]
for i in range(2,5):
    c1.append(fila["A"+str(i)].value)
for i in range(2,5):
    c2.append(fila["B"+str(i)].value)

fisier.create_sheet("PAG2")
fila=fisier["PAG2"]
fila["A1"]="Cheie"
fila["B1"]="Valoare"
for i in range(2,5):
    fila["A"+str(i)]=c1[i-2]
for i in range(2,5):
    fila["B"+str(i)]=random.choice(c2)
fisier.save("Ex.xlsx")


